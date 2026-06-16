#!/usr/bin/env python3
"""Parse Monica's Cobrança Xdental spreadsheet export into receivables forecast JSON.

Input is an XLSX export from Google Sheets. No secrets here.
The sheet is manually edited, so parsing is intentionally tolerant and conservative.
"""
import json
import math
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone, timedelta, date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
IN = ROOT / "data" / "cobranca-xdental.xlsx"
OUT = ROOT / "data" / "receivables-dashboard.json"
REVENUE_JSON = ROOT / "data" / "revenue-dashboard.json"
BRT = timezone(timedelta(hours=-3))
NOW = datetime.now(BRT)
CURRENT_MONTH = NOW.strftime("%Y-%m")

FX_TO_BRL = {"USD": 5.35, "BRL": 1.0, "EUR": 5.85, "MXN": 0.29}
MONTHS = {
    "JANEIRO": 1, "ENERO": 1, "ENE": 1,
    "FEVEREIRO": 2, "FEBRERO": 2, "FEV": 2, "FEB": 2,
    "MARCO": 3, "MARÇO": 3, "MARZO": 3, "MAR": 3,
    "ABRIL": 4, "ABR": 4,
    "MAIO": 5, "MAYO": 5, "MAY": 5,
    "JUNHO": 6, "JHULO": 7, "JULHO": 7, "JULIO": 7, "JUL": 7,
    "AGOSTO": 8, "AGO": 8,
    "SETEMBRO": 9, "SEPTIEMBRE": 9, "SEP": 9,
    "OUTUBRO": 10, "OCTUBRE": 10, "OCT": 10,
    "NOVEMBRO": 11, "NOVIEMBRE": 11, "NOV": 11,
    "DEZEMBRO": 12, "DEXEMBRO": 12, "DICIEMBRE": 12, "DIC": 12, "DEC": 12,
}
CANCEL_WORDS = ("cancel", "reembolso", "refund", "contest", "chargeback")
RISK_WORDS = ("bloque", "cobran", "pend", "debe", "deuda", "atras", "vencid", "no va", "não vai", "nao vai", "solicitei")


def clean(s):
    return re.sub(r"\s+", " ", str(s or "")).strip()


def norm(s):
    return clean(s).upper().replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U").replace("Ç", "C")


def to_int(v):
    try:
        if v is None or v == "":
            return None
        return int(float(str(v).replace(",", ".")))
    except Exception:
        return None


def parse_money(v, default_currency="USD"):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if math.isnan(v):
            return None
        return {"amount": float(v), "currency": default_currency}
    s = clean(v)
    low = s.lower()
    if any(w in low for w in CANCEL_WORDS):
        return {"status": "cancelled"}
    # Ignore operational text without digits.
    if not re.search(r"\d", s):
        return None
    currency = default_currency
    if "R$" in s.upper() or re.search(r"\bRS\b", s.upper()):
        currency = "BRL"
    elif "$" in s:
        currency = "USD"
    # Fix common typos: RS 123, S540,25, $ 190.18
    raw = re.sub(r"[^0-9,.-]", "", s)
    if not raw:
        return None
    # If both separators exist, decide by last separator.
    if "," in raw and "." in raw:
        if raw.rfind(",") > raw.rfind("."):
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(",", "")
    elif "," in raw:
        raw = raw.replace(".", "").replace(",", ".")
    try:
        amount = float(raw)
    except Exception:
        return None
    return {"amount": amount, "currency": currency}


def brl(amount, currency):
    return float(amount or 0) * FX_TO_BRL.get((currency or "USD").upper(), 1.0)


def sale_year(row_date):
    if isinstance(row_date, datetime):
        return row_date.year
    txt = clean(row_date)
    m = re.search(r"20\d{2}", txt)
    if m:
        return int(m.group(0))
    # Launch/current sheets with textual dates and no year are 2026 in this workbook.
    return 2026


def sale_datetime(row_date, sheet_title=""):
    """Parse purchase date; text rows without year are inferred from workbook context.

    Monica's operational rule uses purchase date as D0. Installments are due every 30 days
    from this date, regardless of the month label written in the CUOTA header.
    """
    if isinstance(row_date, datetime):
        return row_date
    if isinstance(row_date, date):
        return datetime(row_date.year, row_date.month, row_date.day)
    txt = clean(row_date).lower()
    if not txt:
        return None
    # ISO-ish fallback: 2026-04-15 22:23:05 or 15/04/2026
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(txt[:19], fmt)
        except Exception:
            pass
    txt_norm = norm(txt).lower().replace(".", " ").replace(" de ", " ")
    parts = txt_norm.split()
    day = None; month = None; year = None
    for part in parts:
        if day is None and part.isdigit() and 1 <= int(part) <= 31:
            day = int(part); continue
        if part.isdigit() and len(part) == 4 and part.startswith("20"):
            year = int(part); continue
        up = norm(part)
        if up in MONTHS:
            month = MONTHS[up]
    if day and month:
        if year is None:
            sheet_norm = norm(sheet_title)
            # Black Ortho rows are from Nov/Dec 2025; other no-year rows in current sheets are 2026.
            year = 2025 if "BLACK ORTHO" in sheet_norm and month in (11, 12) else 2026
        try:
            return datetime(year, month, day)
        except Exception:
            return None
    return None


def due_month_from_purchase(row_date, quota_index, sheet_title=""):
    dt = sale_datetime(row_date, sheet_title)
    if not dt:
        # Conservative fallback only when purchase date is unreadable.
        return header_month("", row_date, quota_index, sheet_title)
    due = dt + timedelta(days=30 * max(0, quota_index - 1))
    return due.strftime("%Y-%m")


def cell_rgb(cell):
    try:
        fg = cell.fill.fgColor
        if fg.type == "rgb" and fg.rgb:
            return fg.rgb.upper()
    except Exception:
        pass
    return ""


def color_status(cell):
    """Interpret Monica's color convention.

    Green: received. Orange/yellow/bright: overdue/defaulted. Red: refund/cancelled.
    """
    rgb = cell_rgb(cell)
    if not rgb or rgb in ("00000000", "FFFFFFFF"):
        return None
    if rgb in ("FFD9EAD3", "FFB6D7A8", "FF93C47D"):
        return "received"
    if rgb in ("FFFF0000", "FFCC0000", "FFEA9999", "FFF4CCCC"):
        return "cancelled"
    if rgb in ("FFFF9900", "FFFFC000", "FFFFD966", "FFFFE599", "FFFF00FF", "FFFF6D01"):
        return "overdue"
    return "marked"


def add_month(year, month, offset):
    idx = (year * 12 + (month - 1)) + offset
    return f"{idx // 12:04d}-{idx % 12 + 1:02d}"


def header_month(header, row_date, quota_index, sheet_title=""):
    h = norm(header)
    for name, month in MONTHS.items():
        if name in h:
            # If a sheet started in Nov/Dec and current installments cross year, infer year from sale.
            y = sale_year(row_date)
            sheet_norm = norm(sheet_title)
            if "BLACK ORTHO" in sheet_norm:
                y = 2025 if month in (11, 12) else 2026
            # For old Nov/Dec headers followed by Jan/Feb, year rolls forward.
            elif month < 4 and any(x in h for x in ["JANEIRO", "FEVEREIRO", "MARCO", "MARÇO", "ABRIL"]) and y >= 2025:
                # If sale month in Nov/Dec, Jan+ belongs next year.
                dt = row_date if isinstance(row_date, datetime) else None
                if dt and dt.month >= 9 and month <= 4:
                    y += 1
            return f"{y:04d}-{month:02d}"
    # Fallback: infer from sale date + quota index - 1.
    if isinstance(row_date, datetime):
        return add_month(row_date.year, row_date.month, max(0, quota_index - 1))
    return None


def find_header(ws):
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=12, max_col=60, values_only=True), 1):
        vals = [clean(v) for v in row]
        joined = " | ".join(vals).upper()
        if "PLATAFORMA" in joined and "CUOTA" in joined:
            return r, vals
    return None, []


def product_from_sheet(sheet, row, idx):
    # Some sheets have PRODUCTO column, others product is the sheet title.
    if "producto" in idx:
        val = clean(row[idx["producto"]])
        if val:
            return val
    return sheet


def parse_workbook(path=IN):
    # read_only=False is intentional: Monica uses cell fill colors as financial status.
    wb = load_workbook(path, data_only=True, read_only=False)
    installment_rows = []
    people_risk = set()
    cancelled_rows = 0
    parsed_rows = 0
    sheet_stats = []

    for ws in wb.worksheets:
        header_row, headers = find_header(ws)
        if not header_row:
            continue
        idx = {}
        for i, h in enumerate(headers):
            hn = norm(h)
            if hn.startswith("PLATAFORMA") or hn == "PLATAFORMA DE VENDA": idx["platform"] = i
            elif "DATA" in hn and ("VENDA" in hn or "COMPRA" in hn): idx["date"] = i
            elif "PRODUCTO" in hn or "PRODUTO" in hn: idx["producto"] = i
            elif "VALOR PLANO" in hn: idx["plan"] = i
            elif "NUMERO DE PARCELAS" in hn or "QNTD" in hn: idx["installments"] = i
            elif "NOME DO USUARIO" in hn: idx["name"] = i
            elif "E-MAIL" in hn or "EMAIL" in hn: idx["email"] = i
            elif "ESTADO" in hn: idx["status"] = i
            elif "OBSERV" in hn: idx["obs"] = i
            elif "COBRAN" in hn: idx["cobranca"] = i
        quota_cols = []
        for i, h in enumerate(headers):
            hn = norm(h)
            if "CUOTA" in hn and ("HOTMART" in hn or "STRIPE" in hn):
                m = re.search(r"CUOTA\s*(\d+)?", hn)
                qi = int(m.group(1)) if m and m.group(1) else len([q for q in quota_cols if ("STRIPE" in hn) == ("STRIPE" in norm(q[1]))]) + 1
                platform = "Stripe" if "STRIPE" in hn else "Hotmart"
                quota_cols.append((i, h, qi, platform))
        sheet_rows = 0
        for cells in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=False):
            cell_row = list(cells) + [None] * 80
            row = [c.value if c is not None else None for c in cell_row]
            platform = clean(row[idx.get("platform", 0)]).title()
            if platform not in ("Stripe", "Hotmart"):
                continue
            ninst = to_int(row[idx.get("installments", -1)]) or 1
            if ninst <= 0:
                continue
            parsed_rows += 1; sheet_rows += 1
            row_date = row[idx.get("date", -1)] if "date" in idx else None
            product = product_from_sheet(ws.title, row, idx)
            name = clean(row[idx.get("name", -1)]) if "name" in idx else ""
            email = clean(row[idx.get("email", -1)]).lower() if "email" in idx else ""
            status_text = " ".join(clean(row[idx[k]]) for k in ("status", "obs", "cobranca") if k in idx).lower()
            row_cancelled = any(w in status_text for w in CANCEL_WORDS)
            row_risk = any(w in status_text for w in RISK_WORDS)
            if row_risk:
                people_risk.add(email or name)
            plan = parse_money(row[idx.get("plan", -1)], "USD") if "plan" in idx else None
            row_values = []
            for col, h, qi, q_platform in quota_cols:
                if q_platform != platform or qi > ninst:
                    continue
                default_cur = "BRL" if platform == "Stripe" else "USD"
                parsed = parse_money(row[col], default_cur)
                if parsed and parsed.get("amount") is not None:
                    row_values.append(parsed)
            if row_values:
                avg_brl = sum(brl(v["amount"], v["currency"]) for v in row_values) / len(row_values)
            elif plan and plan.get("amount") is not None:
                avg_brl = brl(plan["amount"] / max(1, ninst), plan["currency"])
            else:
                avg_brl = 0
            for col, h, qi, q_platform in quota_cols:
                if q_platform != platform or qi > ninst:
                    continue
                due_month = due_month_from_purchase(row_date, qi, ws.title)
                if not due_month:
                    continue
                parsed = parse_money(row[col], "BRL" if platform == "Stripe" else "USD")
                cstatus = color_status(cell_row[col])
                status = "received"
                amount_brl = 0.0
                if row_cancelled or cstatus == "cancelled" or (parsed and parsed.get("status") == "cancelled"):
                    status = "cancelled"; cancelled_rows += 1
                elif cstatus == "received":
                    amount_brl = brl(parsed["amount"], parsed["currency"]) if parsed and parsed.get("amount") is not None else avg_brl
                    status = "received"
                elif cstatus in ("overdue", "marked"):
                    amount_brl = brl(parsed["amount"], parsed["currency"]) if parsed and parsed.get("amount") is not None else avg_brl
                    status = "overdue"
                elif due_month > CURRENT_MONTH:
                    amount_brl = brl(parsed["amount"], parsed["currency"]) if parsed and parsed.get("amount") is not None else avg_brl
                    status = "pending"
                elif parsed and parsed.get("amount") is not None:
                    amount_brl = brl(parsed["amount"], parsed["currency"])
                    status = "received"
                else:
                    amount_brl = avg_brl
                    status = "overdue" if due_month < CURRENT_MONTH else "pending"
                installment_rows.append({
                    "sheet": ws.title,
                    "product": product,
                    "platform": platform,
                    "month": due_month,
                    "quota": qi,
                    "status": status,
                    "amount_brl_est": round(amount_brl, 2),
                    "name": name,
                    "email": email,
                    "risk": bool(row_risk),
                })
        sheet_stats.append({"sheet": ws.title, "rows": sheet_rows, "quota_cols": len(quota_cols)})

    by_month = defaultdict(lambda: {"received_brl": 0.0, "expected_brl": 0.0, "overdue_brl": 0.0, "pending_installments": 0, "overdue_installments": 0})
    by_product = defaultdict(lambda: {"received_brl": 0.0, "expected_brl": 0.0, "overdue_brl": 0.0, "pending_installments": 0, "overdue_installments": 0})
    sample_pending = []
    for it in installment_rows:
        m = by_month[it["month"]]
        p = by_product[it["product"]]
        if it["status"] == "received":
            m["received_brl"] += it["amount_brl_est"]; p["received_brl"] += it["amount_brl_est"]
        elif it["status"] == "pending":
            m["expected_brl"] += it["amount_brl_est"]; p["expected_brl"] += it["amount_brl_est"]
            m["pending_installments"] += 1; p["pending_installments"] += 1
            if len(sample_pending) < 30: sample_pending.append(it)
        elif it["status"] == "overdue":
            m["overdue_brl"] += it["amount_brl_est"]; p["overdue_brl"] += it["amount_brl_est"]
            m["overdue_installments"] += 1; p["overdue_installments"] += 1
            if len(sample_pending) < 30: sample_pending.append(it)

    def finish_map(d):
        out = []
        for name, v in d.items():
            row = {"name": name, **{k: round(val, 2) if isinstance(val, float) else val for k, val in v.items()}}
            out.append(row)
        return out

    months = sorted(finish_map(by_month), key=lambda x: x["name"])
    products = sorted(finish_map(by_product), key=lambda x: (x.get("expected_brl", 0) + x.get("overdue_brl", 0)), reverse=True)
    future = [m for m in months if m["name"] >= CURRENT_MONTH]
    overdue = [m for m in months if m["name"] < CURRENT_MONTH and m.get("overdue_brl", 0) > 0]
    payload = {
        "generated_at": NOW.isoformat(),
        "timezone": "America/Sao_Paulo",
        "source": "Google Sheets export — Cobrança Xdental",
        "source_file": str(path),
        "spreadsheet_id": "1psxy_r11E6g03TudqWBVWjPTJJZOiux-9TsZZHYeYSA",
        "status": "beta",
        "current_month": CURRENT_MONTH,
        "summary": {
            "rows": parsed_rows,
            "installments": len(installment_rows),
            "future_expected_brl": round(sum(m.get("expected_brl", 0) for m in future), 2),
            "overdue_brl": round(sum(m.get("overdue_brl", 0) for m in overdue), 2),
            "pending_installments": int(sum(m.get("pending_installments", 0) for m in future)),
            "overdue_installments": int(sum(m.get("overdue_installments", 0) for m in overdue)),
            "at_risk_count": len([x for x in people_risk if x]),
            "cancelled_cells": cancelled_rows,
        },
        "months": months,
        "future_months": future,
        "products": products[:30],
        "sheet_stats": sheet_stats,
        "sample_pending": sample_pending,
        "notes": [
            "Valores BRL são estimados. Hotmart sem símbolo é tratado como USD; Stripe sem símbolo é tratado como BRL.",
            "Regra Monica/Ruan: vencimento de cada cota = data da compra + 30 dias × (nº da cota - 1); mês do cabeçalho é apenas referência visual.",
            "Células vazias dentro do número de parcelas viram pendente ou vencido conforme vencimento calculado por D0+30 dias.",
            "Linhas/células com Cancelado/Reembolso/Contestação são removidas da previsão e marcadas como canceladas.",
        ],
    }
    return payload


def main():
    payload = parse_workbook()
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    # Merge compact view into dashboard JSON.
    if REVENUE_JSON.exists():
        data = json.loads(REVENUE_JSON.read_text(encoding="utf-8"))
        s = payload["summary"]
        data["receivables"] = {
            "status": "beta",
            "source": payload["source"],
            "spreadsheet_id": payload["spreadsheet_id"],
            "spreadsheet_name": "Cobrança Xdental",
            "generated_at": payload["generated_at"],
            "months": [{"month": m["name"], "expected_brl": m.get("expected_brl", 0), "overdue_brl": m.get("overdue_brl", 0), "received_brl": m.get("received_brl", 0), "pending_installments": m.get("pending_installments", 0), "overdue_installments": m.get("overdue_installments", 0)} for m in payload["future_months"][:12]],
            "products": payload["products"][:12],
            "pending_brl": s["future_expected_brl"],
            "overdue_brl": s["overdue_brl"],
            "pending_installments": s["pending_installments"],
            "overdue_installments": s["overdue_installments"],
            "at_risk_count": s["at_risk_count"],
            "source_note": "Forecast beta gerado da planilha Cobrança Xdental pela regra D0+30 dias e cores da Monica.",
        }
        REVENUE_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"out": str(OUT), "summary": payload["summary"], "future_months": payload["future_months"][:6]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
